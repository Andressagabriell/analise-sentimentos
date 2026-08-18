"""
excel_batch.py
--------------
Processa uma planilha Excel em lote: lê uma coluna de textos, analisa o
sentimento de cada linha e gera uma nova planilha com os resultados.

Uso:
    python excel_batch.py entrada.xlsx
    python excel_batch.py entrada.xlsx --coluna "Comentario" --saida resultado.xlsx

A planilha de entrada precisa ter pelo menos uma coluna com os textos a
analisar (por padrão, o script procura uma coluna chamada "texto",
"comentario", "comentário", "review" ou "mensagem" - case-insensitive -
ou você pode indicar explicitamente com --coluna).
"""

import argparse
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from sentiment_core import analisar_lote

COLUNAS_CANDIDATAS = ["texto", "comentario", "comentário", "review", "mensagem", "feedback"]

CORES_PREENCHIMENTO = {
    "Positivo": "C6EFCE",  # verde claro
    "Negativo": "FFC7CE",  # vermelho claro
    "Neutro": "F2F2F2",    # cinza claro
}
CORES_FONTE = {
    "Positivo": "006100",
    "Negativo": "9C0006",
    "Neutro": "606060",
}


def detectar_coluna(df: pd.DataFrame) -> str:
    """Tenta descobrir automaticamente qual coluna contém os textos."""
    colunas_lower = {c.lower().strip(): c for c in df.columns}
    for candidata in COLUNAS_CANDIDATAS:
        if candidata in colunas_lower:
            return colunas_lower[candidata]
    raise ValueError(
        f"Não encontrei automaticamente a coluna de texto. "
        f"Colunas disponíveis: {list(df.columns)}. "
        f"Use --coluna \"NomeDaColuna\" para indicar manualmente."
    )


def _ler_tabela(caminho_entrada: str) -> pd.DataFrame:
    """Lê a planilha de entrada, aceitando tanto .xlsx quanto .csv."""
    if caminho_entrada.lower().endswith(".csv"):
        # tenta detectar automaticamente o separador (vírgula ou ponto-e-vírgula,
        # comum em CSVs exportados em português/Excel-BR)
        try:
            return pd.read_csv(caminho_entrada, sep=None, engine="python")
        except Exception as e:
            raise ValueError(f"Não consegui ler o CSV: {e}")
    else:
        try:
            return pd.read_excel(caminho_entrada)
        except ValueError as e:
            raise ValueError(
                f"Não consegui abrir '{caminho_entrada}' como planilha Excel "
                f"(erro: {e}). Se o arquivo foi salvo como texto/CSV com a "
                f"extensão .xlsx errada, renomeie para .csv e rode de novo."
            )


def processar_dataframe(df: pd.DataFrame, coluna: str | None = None) -> pd.DataFrame:
    """
    Núcleo do processamento em lote, sem depender de arquivos em disco.
    Recebe um DataFrame já carregado e devolve outro com as colunas de
    análise de sentimento adicionadas.

    Reutilizada tanto pelo CLI (processar_planilha) quanto pela interface
    web (app.py, via upload de arquivo).
    """
    if df.empty:
        raise ValueError("A planilha de entrada está vazia.")

    if coluna is None:
        coluna = detectar_coluna(df)
    elif coluna not in df.columns:
        raise ValueError(f"Coluna '{coluna}' não encontrada. Colunas disponíveis: {list(df.columns)}")

    textos = df[coluna].fillna("").astype(str).tolist()
    resultados = analisar_lote(textos)

    df_resultado = df.copy()
    df_resultado["Sentimento"] = [r["classificacao"] for r in resultados]
    df_resultado["Score (compound)"] = [r["compound"] for r in resultados]
    df_resultado["Negativo (%)"] = [r["neg"] for r in resultados]
    df_resultado["Neutro (%)"] = [r["neu"] for r in resultados]
    df_resultado["Positivo (%)"] = [r["pos"] for r in resultados]
    df_resultado["Texto traduzido (EN)"] = [r["texto_traduzido"] for r in resultados]

    return df_resultado


def processar_planilha(caminho_entrada: str, coluna: str | None, caminho_saida: str) -> str:
    df = _ler_tabela(caminho_entrada)
    print(f"Total de linhas a processar: {len(df)}")

    coluna_usada = coluna if coluna is not None else detectar_coluna(df)
    print(f"Coluna de texto identificada: '{coluna_usada}'")

    df_resultado = processar_dataframe(df, coluna_usada)
    gerar_planilha_formatada(df_resultado, caminho_saida)

    n_erros = sum(1 for c in df_resultado["Sentimento"] if str(c).startswith("ERRO"))
    if n_erros:
        print(f"⚠️  {n_erros} linha(s) tiveram erro de tradução/análise (ver coluna 'Sentimento').")

    return caminho_saida


def gerar_planilha_formatada(df: pd.DataFrame, destino) -> None:
    """
    Escreve o DataFrame já analisado como .xlsx formatado (cabeçalho colorido,
    linhas coloridas por sentimento, colunas com largura automática).

    `destino` pode ser um caminho de arquivo (str) ou um buffer em memória
    (io.BytesIO) — este último é usado pela interface web (app.py), que
    processa o upload sem precisar gravar nada em disco.
    """
    from openpyxl import load_workbook

    df.to_excel(destino, index=False, sheet_name="Análise de Sentimento")

    # Se for um buffer em memória, precisa voltar o cursor para o início
    # antes de reabrir com load_workbook.
    if hasattr(destino, "seek"):
        destino.seek(0)

    wb = load_workbook(destino)
    ws = wb.active

    fonte_padrao = "Arial"
    idx_sentimento = list(df.columns).index("Sentimento") + 1  # 1-indexed

    # Cabeçalho
    for cell in ws[1]:
        cell.font = Font(name=fonte_padrao, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas de dados
    for row_idx in range(2, ws.max_row + 1):
        sentimento = ws.cell(row=row_idx, column=idx_sentimento).value
        cor_fundo = CORES_PREENCHIMENTO.get(sentimento, "FFFFFF")
        cor_fonte = CORES_FONTE.get(sentimento, "000000")

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name=fonte_padrao, color=cor_fonte if col_idx == idx_sentimento else "000000")
            if col_idx == idx_sentimento:
                cell.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")

    # Largura automática (aproximada) das colunas
    for col_idx, col_name in enumerate(df.columns, start=1):
        letra = get_column_letter(col_idx)
        maior_valor = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()[:200]]
        )
        ws.column_dimensions[letra].width = min(max(maior_valor + 2, 12), 50)

    ws.freeze_panes = "A2"  # congela o cabeçalho

    if hasattr(destino, "seek"):
        destino.seek(0)
        destino.truncate()  # remove qualquer byte residual do save anterior
        wb.save(destino)
        destino.seek(0)
    else:
        wb.save(destino)


def main():
    parser = argparse.ArgumentParser(description="Analisa sentimento de uma planilha em lote (.xlsx ou .csv).")
    parser.add_argument("entrada", help="Caminho da planilha de entrada (.xlsx ou .csv)")
    parser.add_argument("--coluna", default=None, help="Nome da coluna com os textos (opcional, autodetecta)")
    parser.add_argument("--saida", default=None, help="Caminho da planilha de saída (opcional, sempre .xlsx)")
    args = parser.parse_args()

    if args.saida:
        caminho_saida = args.saida
    else:
        base = args.entrada
        for ext in (".xlsx", ".csv", ".xls"):
            if base.lower().endswith(ext):
                base = base[: -len(ext)]
                break
        caminho_saida = f"{base}_analisado.xlsx"

    try:
        resultado_path = processar_planilha(args.entrada, args.coluna, caminho_saida)
        print(f"\n✅ Concluído! Planilha gerada em: {resultado_path}")
    except Exception as e:
        print(f"\n❌ Erro: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()